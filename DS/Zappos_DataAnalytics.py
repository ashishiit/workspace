'''
Created on Sep 6, 2017

@author: s528358
'''
import pandas as pd
import matplotlib.pyplot as plt
from xlrd import open_workbook
from openpyxl import Workbook, load_workbook
from pandas import ExcelFile
from pandas import ExcelWriter
import numpy as np

# count_cancelled = 0
# wb = Workbook()
# wb = open_workbook('data_science_analytics_2018_data.xlsx')
# print (wb.sheet_names()[0])
# ws = wb.active
# print (ws.sheetnames)
# wb = load_workbook('data_science_analytics_2018_data.xlsx')
# ws = wb['data']
# count = 0
# print (ws.rows.values())
# print (ws['A1'])

# for row in ws.rows:
#     count = count + 1
#     for cell in row:
#         print (cell.value,end = ' ')
#     print('\n')
# print('total lines = ',count)
# print (ws.get_sheet_names())
# df = pd.read_excel('data_science_analytics_2018_data.xlsx', index_col = 'InvoiceNo')
# print (df.head())
# print(df['UnitPrice'])
# description =  df['Description'].values
# print ('number of columns = ',len(description))
# print (date)
# for i in invoice:
#     print (i)
# print (description[0])
# FORMAT = ['InvoiceNo', 'StockCode', 'Description' 'Quantity', 'InvoiceDate', 'UnitPrice', 'CustomerID', 'Country']
# info = df[FORMAT]
# invoice = df['InvoiceNo'].values
# invoice_str = []
# for i in invoice:
#     invoice_str.append(str(i))
#print(invoice_str)
# for i in invoice_str:
#     if 'C' in i:
#         print (i)
# data = [['one',1],['two', 2],['three', 3],['four', 4]]
# df = pd.DataFrame(data, columns = ['words', 'integer'])
# print(df['words'])
# print (df.iloc[0])
# print (invoice)
# for i in invoice:
#     if 'C' in i:
#         print (i)
df = pd.read_excel('data_science_analytics_2018_data.xlsx', sheetname='data')
# print ('column headings')
# invoice = df['InvoiceNo']
# print (df.index)
list_InvoiceNo = df['InvoiceNo']
list_StockCode = df['StockCode']
list_Description = df['Description']
list_Quantity = df['Quantity']
list_InvoiceDate = df['InvoiceDate']
list_UnitPrice = df['UnitPrice']
list_CustomerID = df['CustomerID']
list_Country = df['Country']
# print ('total InvoiceNo = ',len(list_InvoiceNo))
# print (df['UnitPrice'].describe())
# print ('--------------------------')
# print ('maximum value = ',df['UnitPrice'].max())
# print ('minimum value = ', df['UnitPrice'].min())
# print (df.ix[1])
# print (df[df.duplicated(['InvoiceNo'], keep=False)])
# count = 0

# for i in df.index:
#     if 'C' in str(df['InvoiceNo'][i]):
#         print (list_InvoiceNo[i], list_StockCode[i], list_Description[i], list_Quantity[i], list_InvoiceDate[i], list_UnitPrice[i], list_CustomerID[i], list_Country[i])
#         count = count + 1
# print('number of matches = %d'%count)
#         print (list_InvoiceNo[i], list_StockCode[i], list_Description[i], list_Quantity[i], list_InvoiceDate[i], list_UnitPrice[i], list_CustomerID[i], list_Country[i])
# for i in df.index:
#     print(df['InvoiceNo'][i])
# d = {'one' : np.random.rand(10),
#      'two' : np.random.rand(10)}

# df = pd.DataFrame(['StockCode', 'Quantity'])
# 
# df.plot(style=['StockCode','Quantity'])
# plt.show()
# print (df.index)
# print (df['InvoiceNo'].count())
# print (df.ix[1])
# count_cancelled = 0
def Count_CancelledTransaction():
    count_cancelled = 0
    for i in df.index:
    #     print(i)
        if 'C' in str(df['InvoiceNo'][i]):
            count_cancelled = count_cancelled + 1
    print ('total transaction = ', df['InvoiceNo'].count())
    print ('cancelled transaction = ', count_cancelled)
    print ('valid transaction = ', df['InvoiceNo'].count()-count_cancelled)
Count_CancelledTransaction()
# print ('valid transaction = ', df['InvoiceNo'].count-count_cancelled)